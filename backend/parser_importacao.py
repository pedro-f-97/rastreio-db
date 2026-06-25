from datetime import datetime
from openpyxl import load_workbook
import csv
import io

class ErroParsing(Exception):
    pass

def _parse_valor(raw, separador_decimal: str) -> float:
    if raw is None:
        raise ErroParsing("valor em falta")
    if isinstance(raw, (int, float)):
        return float(raw)
    texto = str(raw).strip().replace(" ", "").replace("\xa0", "")
    if separador_decimal == ",":
        texto = texto.replace(".", "").replace(",", ".")
    else:
        texto = texto.replace(",", "")
    try:
        return float(texto)
    except ValueError:
        raise ErroParsing(f"valor inválido: '{raw}'")

def _parse_data(raw, formato_data: str):
    if raw is None:
        raise ErroParsing("data em falta")
    if hasattr(raw, "date"):
        return raw.date()
    try:
        return datetime.strptime(str(raw).strip(), formato_data).date()
    except ValueError:
        raise ErroParsing(f"data inválida: '{raw}' (formato esperado: {formato_data})")

def _ler_linhas_xlsx(file_bytes: bytes, linha_inicio: int) -> list[tuple]:
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        raise ErroParsing("Ficheiro Excel sem sheets activas.")
    linhas = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i < linha_inicio:
            continue
        linhas.append(row)
    return linhas

def _ler_linhas_csv(file_bytes: bytes, linha_inicio: int, separador_decimal: str) -> list[tuple]:
    texto = file_bytes.decode("utf-8-sig")
    reader = csv.reader(io.StringIO(texto))
    linhas = []
    for i, row in enumerate(reader, start=1):
        if i < linha_inicio:
            continue
        linhas.append(tuple(row))
    return linhas

def parse_ficheiro(file_bytes: bytes, perfil) -> dict:
    """
    Devolve:
    {
        "transacoes": [ {"data": ..., "descricao": ..., "valor": ..., "saldo": ...} ],
        "total_linhas": int,
        "erros": [ {"linha": int, "erro": str} ]
    }
    """
    if perfil.tipo_ficheiro.value == "xlsx":
        linhas = _ler_linhas_xlsx(file_bytes, perfil.linha_inicio_dados)
    else:
        linhas = _ler_linhas_csv(file_bytes, perfil.linha_inicio_dados, perfil.separador_decimal)

    transacoes = []
    erros = []
    numero_linha = perfil.linha_inicio_dados

    for row in linhas:
        numero_linha_atual = numero_linha
        numero_linha += 1

        def celula(indice):
            if indice is None or indice >= len(row):
                return None
            return row[indice]

        # Ignorar linhas completamente vazias
        if all(c is None or str(c).strip() == "" for c in row):
            continue

        try:
            data = _parse_data(celula(perfil.coluna_data), perfil.formato_data)
            descricao = str(celula(perfil.coluna_descricao) or "").strip()

            if perfil.modo_valor.value == "coluna_unica":
                if perfil.coluna_valor is None:
                    raise ErroParsing("coluna_valor não definida para modo coluna_unica")
                valor = _parse_valor(celula(perfil.coluna_valor), perfil.separador_decimal)
            else:
                if perfil.coluna_debito is None or perfil.coluna_credito is None:
                    raise ErroParsing("coluna_debito/coluna_credito não definidas para modo duas_colunas")
                raw_debito = celula(perfil.coluna_debito)
                raw_credito = celula(perfil.coluna_credito)
                debito = _parse_valor(raw_debito, perfil.separador_decimal) if raw_debito not in (None, "", "None") else 0.0
                credito = _parse_valor(raw_credito, perfil.separador_decimal) if raw_credito not in (None, "", "None") else 0.0
                valor = credito - debito

            if perfil.coluna_fee is not None:
                raw_fee = celula(perfil.coluna_fee)
                if raw_fee not in (None, "", "None"):
                    valor += _parse_valor(raw_fee, perfil.separador_decimal)

            saldo = None
            if perfil.tem_saldo and perfil.coluna_saldo is not None:
                raw_saldo = celula(perfil.coluna_saldo)
                if raw_saldo not in (None, "", "None"):
                    saldo = _parse_valor(raw_saldo, perfil.separador_decimal)

            transacoes.append({
                "data": data,
                "descricao": descricao,
                "valor": valor,
                "saldo": saldo,
            })

        except ErroParsing as e:
            erros.append({"linha": numero_linha_atual, "erro": str(e)})

    return {
        "transacoes": transacoes,
        "total_linhas": len(transacoes) + len(erros),
        "erros": erros,
    }