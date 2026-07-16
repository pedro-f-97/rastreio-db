from openpyxl import load_workbook
from database import SessionLocal, Categoria, Subcategoria, Transacao, criar_tabelas
from datetime import datetime

CAMINHO_EXCEL = "../Extrato.xlsx"

MAPEAMENTO_CATEGORIAS = {
    "Salário": "Receita",
    "Reembolso": None,
    "Entretenimento": "Entretenimento",
    "Gasto Doméstico": "Casa",
    "Investimento": "Investimento",
    "Transporte": "Transporte",
    "Saúde": "Saúde",
    "Aparência": "Aparência",
    "Despesa Legal": "Pontual",
}

def normalizar_saldo(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    # Remove espaços e converte vírgula decimal para ponto
    valor = valor.strip().replace(" ", "")
    # Remove pontos de milhares se existir vírgula decimal
    if "," in valor:
        valor = valor.replace(".", "").replace(",", ".")
    return float(valor)

def migrar():
    criar_tabelas()
    session = SessionLocal()
    total_existente = session.query(Transacao).count()
    if total_existente > 0:
        print(f"A base de dados já tem {total_existente} transações. Migração cancelada.")
        session.close()
        return
    session = SessionLocal()

    wb = load_workbook(CAMINHO_EXCEL, read_only=True)
    ws_extratos = wb["Extratos"]
    ws_dados = wb["Dados"]

    # Ler categorias da folha Dados
    categorias_excel = []
    for row in ws_dados.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        categorias_excel.append(row[2])  # coluna Categoria

    # Ler e inserir transações
    transacoes_excel = []
    for row in ws_extratos.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            break
        transacoes_excel.append(row)

    inseridas = 0
    for i, row in enumerate(transacoes_excel):
        if isinstance(row[0], datetime):
            data = row[0]
        elif isinstance(row[0], str):
            data = datetime.strptime(row[0], "%d/%m/%Y")
        else:
            data = None
        descricao = row[2]
        debito = row[3]
        credito = row[4]
        saldo = normalizar_saldo(row[5])

        # Positivo para crédito, negativo para débito
        if credito:
            valor = float(credito)
        elif debito:
            valor = -float(debito)
        else:
            valor = 0.0

        nome_categoria_excel = categorias_excel[i] if i < len(categorias_excel) else None
        nome_categoria_novo = MAPEAMENTO_CATEGORIAS.get(nome_categoria_excel) if nome_categoria_excel else None
        categoria = session.query(Categoria).filter_by(nome=nome_categoria_novo).first() if nome_categoria_novo else None

        t = Transacao(
            data=data,
            descricao=descricao,
            valor=valor,
            saldo=saldo,
            reembolso=nome_categoria_excel == "Reembolso",
            categoria_id=categoria.id if categoria else None,
        )
        session.add(t)
        inseridas += 1

    session.commit()
    session.close()
    print(f"Migração concluída: {inseridas} transações inseridas.")

if __name__ == "__main__":
    migrar()