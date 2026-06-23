from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from database import SessionLocal, PerfilImportacao as PerfilImportacaoModel
from schemas import PerfilImportacaoCreate, PerfilImportacao as PerfilImportacaoSchema
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from openpyxl import load_workbook
import schemas
import io

router = APIRouter(prefix="/perfis-importacao", tags=["perfis-importacao"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.get("/", response_model=list[PerfilImportacaoSchema])
def listar_perfis(db: Session = Depends(get_db)):
    return db.query(PerfilImportacaoModel).order_by(PerfilImportacaoModel.nome).all()

@router.post("/", response_model=PerfilImportacaoSchema, status_code=201)
def criar_perfil(perfil: PerfilImportacaoCreate, db: Session = Depends(get_db)):
    existente = db.query(PerfilImportacaoModel).filter(PerfilImportacaoModel.nome == perfil.nome).first()
    if existente:
        raise HTTPException(status_code=409, detail="Já existe um perfil com esse nome.")
    novo = PerfilImportacaoModel(**perfil.model_dump())
    db.add(novo)
    db.commit()
    db.refresh(novo)
    return novo

@router.post("/analisar-ficheiro")
async def analisar_ficheiro(file: UploadFile = File(...)):
    conteudo = await file.read()
    nome = file.filename or ""
    if nome.endswith(".csv"):
        import csv
        texto = conteudo.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(texto))
        linhas = [[str(c) for c in row] for _, row in zip(range(12), reader)]
    else:
        wb = load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
        if ws is None:
            raise HTTPException(status_code=400, detail="Ficheiro Excel sem sheets activas.")
        linhas = []
        for row in ws.iter_rows(max_row=12, values_only=True):
            linhas.append([str(c) if c is not None else "" for c in row])
    return {"linhas": linhas}

@router.put("/{perfil_id}", response_model=PerfilImportacaoSchema)
def atualizar_perfil(perfil_id: int, dados: PerfilImportacaoCreate, db: Session = Depends(get_db)):
    perfil = db.query(PerfilImportacaoModel).filter(PerfilImportacaoModel.id == perfil_id).first()
    if not perfil:
        raise HTTPException(status_code=404, detail="Perfil não encontrado.")
    conflito = db.query(PerfilImportacaoModel).filter(
        PerfilImportacaoModel.nome == dados.nome,
        PerfilImportacaoModel.id != perfil_id
    ).first()
    if conflito:
        raise HTTPException(status_code=409, detail="Já existe um perfil com esse nome.")
    for campo, valor in dados.model_dump().items():
        setattr(perfil, campo, valor)
    db.commit()
    db.refresh(perfil)
    return perfil

@router.delete("/{perfil_id}", status_code=204)
def eliminar_perfil(perfil_id: int, db: Session = Depends(get_db)):
    perfil = db.query(PerfilImportacaoModel).filter(PerfilImportacaoModel.id == perfil_id).first()
    if not perfil:
        raise HTTPException(status_code=404, detail="Perfil não encontrado.")
    db.delete(perfil)
    db.commit()

@router.patch("/{perfil_id}/conta")
def associar_conta(perfil_id: int, dados: schemas.AssociarConta, db: Session = Depends(get_db)):
    perfil = db.query(PerfilImportacaoModel).filter(PerfilImportacaoModel.id == perfil_id).first()
    if not perfil:
        raise HTTPException(status_code=404, detail="Perfil não encontrado.")
    perfil.conta_id = dados.conta_id
    db.commit()
    db.refresh(perfil)
    return {"ok": True, "conta_id": perfil.conta_id}