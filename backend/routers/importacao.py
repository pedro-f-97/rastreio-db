from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from database import SessionLocal, Transacao, RegraCategorizacao, DB_PATH
from openpyxl import load_workbook
from datetime import datetime
import io
import os
import shutil

router = APIRouter(prefix="/importacao", tags=["importacao"])

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

def aplicar_regras(transacao, regras):
    if transacao.descricao.upper().startswith("TRF"):
        return
    for regra in regras:
        if regra.palavra_chave.upper() in transacao.descricao.upper():
            # Sem categoria — preenche tudo
            if transacao.categoria_id is None:
                transacao.categoria_id = regra.categoria_id
                transacao.subcategoria_id = regra.subcategoria_id
            # Com categoria coincidente mas sem subcategoria — preenche só subcategoria
            elif transacao.categoria_id == regra.categoria_id and transacao.subcategoria_id is None:
                transacao.subcategoria_id = regra.subcategoria_id
            break

@router.post("/")
async def importar_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    # Criar backup automático antes de processar o Excel
    if os.path.exists(DB_PATH):
        shutil.copy2(DB_PATH, DB_PATH + '.pre_import')
    conteudo = await file.read()
    wb = load_workbook(io.BytesIO(conteudo), read_only=True)
    ws = wb.active
    if ws is None:
        raise HTTPException(status_code=400, detail="Ficheiro Excel inválido")

    regras = db.query(RegraCategorizacao).all()

    inseridas = 0
    duplicadas = 0

    for row in ws.iter_rows(min_row=9, values_only=True):
        if row[0] is None:
            continue

        if isinstance(row[0], datetime):
            data = row[0].date()
        elif isinstance(row[0], str):
            data = datetime.strptime(row[0], "%d/%m/%Y").date()
        else:
            continue
        descricao = row[2]
        valor = float(str(row[3])) if row[3] is not None else 0.0
        saldo = float(str(row[4])) if row[4] is not None else 0.0

        # Verificar duplicado
        existente = db.query(Transacao).filter(
            Transacao.data == data,
            Transacao.descricao == descricao,
            Transacao.valor == valor,
            Transacao.saldo == saldo,
        ).first()

        if existente:
            duplicadas += 1
            continue

        nova = Transacao(
            data=data,
            descricao=descricao,
            valor=valor,
            saldo=saldo,
        )
        aplicar_regras(nova, regras)
        db.add(nova)
        inseridas += 1

    db.commit()
    return {"inseridas": inseridas, "duplicadas": duplicadas}